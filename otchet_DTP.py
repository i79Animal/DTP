import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import Font


wb = openpyxl.Workbook()

zalivka = (1, 2, 15, 28, 39, 40, 53, 66, 79, 92, 93, 106, 119, 132, 145, 158, 171, 184)
str_group = [28, 2, 15, 40, 53, 66, 79, 93, 106, 119, 132, 158, 171, 184] # строка = группа
input_data =[[462, 446, 16],                # звонки 28
             [1029, 1057, 110, 24, 4, 0],   # 1 линия, инженеры-консультанты 2
             [156, 162, 7, 0, 0, 0],        # операторы 15
             [41, 43, 4, 1, 1, 1],          # 2 линия, инженеры 40
             [31, 30, 15, 0, 2, 2],         # выездные иненеры 53
             [50, 58, 9, 4, 1, 1],          # 3 линия 66
             [27, 42, 41, 1, 0, 0],         # Учет ИТ активов 79
             [356, 375, 76, 6, 6, 3],       # Отдел поддержки приложений, сервисная группа 93
             [375, 351, 68, 1, 7, 2],       # технологическая группа 106
             [78, 70, 111, 0, 9, 8],        # финансовая группа 119
             [44, 45, 67, 4, 6, 6],         # группа сопровождения сервисных систем 132
             [33, 33, 22, 1, 1, 0],         # аналитическая группа 158
             [11, 9, 6, 1, 3, 1]            # группа по интеграции 171
             ]#[0, 0, 0, 0, 0, 0, 0]]       # добавится общий итог по ДТП 184

def main():
    data_last_week(input_data) # обработка новых данных.
    make_new_file() # создаем файл с объединенными данными

def data_last_week(data):
    data[0].append(round(data[0][2] / data[0][1] * 100, 2)) # добавляем % необработанных звонков
    data.append([0, 0, 0, 0, 0, 0, 0])      # добавляем общий итог по ДТП
    for i in range(1, len(data) - 1):       # добавляем 7 значение(Просроченные -
        data[i].append(data[i][-2] - data[i][-1])
        for j in range(7):
            data[-1][j] += data[i][j]       # - открытые за текущий период)
    print(*data)

def make_new_file():
    # загружаем старый файл и создаем новый
    last_file = openpyxl.load_workbook('отчет ДТП.xlsx', data_only=True)
    sheet = last_file['сводная']
    rows = sheet.max_row
    cols = sheet.max_column
    print(cols)
    wb.create_sheet(title='сводная', index=0)
    wb.remove(wb['Sheet'])
    sheet_new = wb['сводная']
    sheet_new.column_dimensions['A'].width = 40
    sheet_new.freeze_panes = 'B1'
    make_style(wb)
    for i in range(1, rows + 1):
        for j in range(1, cols + 1):
            cell_last_file = sheet.cell(row=i, column=j) # считываем значение ячейки
            if not cell_last_file.value and j == 1: # первая ячейка пустая, тогда пропускаем
                pass # попробовать сделать чтобы пропускать 5 строк
            elif cell_last_file.value: # ячейка не пустая и не равна нулю
                cell_new_file = sheet_new.cell(row=i, column=j) # копируем значение
                cell_new_file.value = cell_last_file.value
            elif cell_last_file.value == 0: # значение ячейки равно нулю
                cell_new_file = sheet_new.cell(row=i, column=j)
                cell_new_file.value = 0 # принудительно копируем ноль
            if i in zalivka: sheet_new.cell(i, j).style = 'Grey' # заливаем нужные строки
            if i == 32 and j != 1:
                cell_new_file.value = round(cell_new_file.value, 2) # 32 округляем
    add_data(wb, str_group, input_data) # вносим новые данные в файл

def add_data(wb, str_group, data):
    sheet = wb['сводная']
    cols = sheet.max_column
    today = datetime.today() - timedelta(days=5)
    monday = today - timedelta(days=6)
    sunday = today
    period = monday.strftime("%d.%m") + '-' + sunday.strftime("%d.%m")
    print(period)
    for i in range(len(str_group)):
        sheet.cell(row=str_group[i], column=cols).value = period # обозначаем неделю
        for j in range(len(data[i])):
            sheet.cell(row=str_group[i]+1+j, column=cols).value = data[i][j]

def make_style(wb): # создаем именованный стиль для серой заливки:
    from openpyxl.styles import NamedStyle, Border, Side, PatternFill
    ns = NamedStyle(name='Grey')
    ns.fill = PatternFill("solid", fgColor="909090")
    border = Side()
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    wb.add_named_style(ns)

def make_sheet(wb, name_sheet, start_row, title=''):
    wb.create_sheet(title=name_sheet)
    sheet = wb[name_sheet]
    sheet.column_dimensions['A'].width = 40
    sheet.freeze_panes = 'B1'
    cols = wb['сводная'].max_column + 1
    print(cols)
    row_delta = 1
    if title != '':  # закрашиваем строку если есть титул
        sheet.cell(row=1, column=1).value = title
        sheet.cell(row=1, column=1).style = 'Grey'
        for col in range(2, 10):
            cell = sheet.cell(row=1, column=col)
            cell.style = 'Grey'
        row_delta = 2
    for row in range(0, 9): # заполняем значения
        value = wb['сводная'].cell(row=start_row + row, column=1).value
        cell = sheet.cell(row=row + row_delta, column=1)
        cell.value = value
        if row == 0: cell.style = 'Grey'
        for col in range(2, 10):
            value = wb['сводная'].cell(row=start_row+row, column=cols-10+col).value
            cell = sheet.cell(row=row+row_delta, column=col)
            cell.value = value
            if row == 0: cell.style = 'Grey' # заливаем нужные строки
    # далее переделываем в получение данных списком.

main()
# создаем остальные листы и копируем в них данные
# лист ДТП
make_sheet(wb, 'ДТП', 184)
make_sheet(wb, '3 линия', 66)
make_sheet(wb, '1 линия', 2, '1 линия')
#make_sheet('Учет ИТ активов', 79)
wb.save('result.xlsx')